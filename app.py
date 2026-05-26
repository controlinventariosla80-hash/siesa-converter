"""
Convertidor de informes SIESA (archivos planos) a Excel.
Corre localmente con: streamlit run app.py
"""

import io
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import siesa_parser

# ---------------------------------------------------------------------------
# Configuración de página
# ---------------------------------------------------------------------------

st.set_page_config(
    page_title="SIESA → Excel",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.title("📊 Convertidor SIESA → Excel")
st.markdown(
    "Sube un archivo plano de informe SIESA (`.txt`, `.prn`, `.psa`, `.rpt`) "
    "y descarga un Excel limpio con los datos listos para analizar."
)
st.divider()

# ---------------------------------------------------------------------------
# Carga del archivo
# ---------------------------------------------------------------------------

uploaded = st.file_uploader(
    "Selecciona el archivo de informe SIESA",
    type=["txt", "prn", "psa", "rpt", "TXT", "PRN", "PSA", "RPT"],
    help="El archivo puede estar en cualquier codificación (UTF-8, Latin-1, etc.)",
)

if not uploaded:
    st.info("⬆️ Sube un archivo para comenzar.")
    st.stop()

# ---------------------------------------------------------------------------
# Procesamiento
# ---------------------------------------------------------------------------

with st.spinner("Leyendo y procesando el archivo..."):
    raw = uploaded.read()
    report = siesa_parser.parse(raw)

if report.dataframe is None or report.dataframe.empty:
    st.error(
        "❌ No se pudo detectar la estructura de datos del archivo. "
        "Verifica que sea un informe SIESA válido con encabezados de columna separados por líneas de guiones."
    )
    st.stop()

st.success(f"✅ Archivo procesado — {len(report.dataframe):,} filas · {len(report.columns)} columnas")

# ---------------------------------------------------------------------------
# Diagnóstico (siempre visible para facilitar ajustes)
# ---------------------------------------------------------------------------

with st.expander("🔍 Diagnóstico de columnas detectadas", expanded=(len(report.columns) <= 3)):
    d = report.debug
    if d:
        st.markdown(f"**Separadores encontrados en el archivo:** {d.get('sep_indices', [])} ...")
        st.markdown(f"**Sección de columnas:** entre los separadores número {d.get('col_h_sep_before')} y {d.get('col_h_sep_after')} (en la lista anterior)")
        st.markdown("**Líneas de encabezado de columna leídas:**")
        for ln in d.get("col_header_lines_raw", []):
            st.code(ln)
        st.markdown("**Posiciones y nombres de columna detectados:**")
        for pos, name in d.get("col_positions", []):
            st.markdown(f"- Posición `{pos}` → **{name}**")
    else:
        st.info("Sin datos de diagnóstico.")

# ---------------------------------------------------------------------------
# Metadata del informe
# ---------------------------------------------------------------------------

with st.expander("ℹ️ Información del informe", expanded=True):
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Código", report.report_code or "—")
    col2.metric("Fecha", report.report_date or "—")
    col3.metric("Empresa", report.company or "—")
    col4.metric("Título", report.report_title or "—")

    if report.filters:
        st.markdown("**Filtros aplicados:**")
        filter_cols = st.columns(min(len(report.filters), 3))
        for i, (k, v) in enumerate(report.filters.items()):
            filter_cols[i % 3].markdown(f"- **{k}:** {v}")

# ---------------------------------------------------------------------------
# Vista previa de datos
# ---------------------------------------------------------------------------

st.subheader("Vista previa")
st.dataframe(report.dataframe, use_container_width=True, height=400)

# ---------------------------------------------------------------------------
# Generación del Excel
# ---------------------------------------------------------------------------

def _build_excel(report: siesa_parser.SiesaReport) -> bytes:
    wb = openpyxl.Workbook()

    # ---- Hoja principal: Datos ----
    ws_data = wb.active
    ws_data.title = "Datos"

    # Estilo de encabezado
    HDR_FILL = PatternFill("solid", fgColor="1F497D")
    HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # Encabezados
    for col_idx, col_name in enumerate(report.columns, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = HDR_ALIGN
        cell.border = BORDER

    ws_data.row_dimensions[1].height = 30

    # Detectar columnas numéricas
    numeric_cols = set()
    for col_name in report.columns:
        if col_name in report.dataframe.columns:
            series = report.dataframe[col_name].dropna()
            if len(series) > 0:
                numeric_count = series.apply(lambda x: isinstance(x, (int, float))).sum()
                if numeric_count / len(series) > 0.5:
                    numeric_cols.add(col_name)

    # Datos
    EVEN_FILL = PatternFill("solid", fgColor="F2F2F2")
    DATA_FONT = Font(name="Calibri", size=10)

    for row_idx, row_tuple in enumerate(report.dataframe.itertuples(index=False), 2):
        is_even = (row_idx % 2 == 0)
        for col_idx, col_name in enumerate(report.columns, 1):
            value = row_tuple[col_idx - 1]
            cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = BORDER

            if is_even:
                cell.fill = EVEN_FILL

            if col_name in numeric_cols and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Ajuste automático de ancho de columnas
    for col_idx, col_name in enumerate(report.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        for row_idx in range(2, ws_data.max_row + 1):
            val = ws_data.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws_data.column_dimensions[col_letter].width = min(max_len + 4, 55)

    # Congelar fila de encabezado
    ws_data.freeze_panes = "A2"

    # Autofiltro
    ws_data.auto_filter.ref = ws_data.dimensions

    # ---- Hoja secundaria: Información del informe ----
    ws_info = wb.create_sheet("Información")
    INFO_FONT_TITLE = Font(bold=True, name="Calibri", size=12, color="1F497D")
    INFO_FONT = Font(name="Calibri", size=10)
    INFO_LABEL_FONT = Font(bold=True, name="Calibri", size=10)

    ws_info.column_dimensions["A"].width = 25
    ws_info.column_dimensions["B"].width = 50

    info_rows = [
        ("Código de informe", report.report_code),
        ("Título", report.report_title),
        ("Empresa", report.company),
        ("Fecha del informe", report.report_date),
        ("Archivo origen", ""),
        ("Filas procesadas", len(report.dataframe)),
        ("Columnas", len(report.columns)),
    ]

    ws_info.cell(row=1, column=1, value="Información del Informe SIESA").font = INFO_FONT_TITLE
    ws_info.merge_cells("A1:B1")

    for r, (label, value) in enumerate(info_rows, 3):
        ws_info.cell(row=r, column=1, value=label).font = INFO_LABEL_FONT
        ws_info.cell(row=r, column=2, value=value).font = INFO_FONT

    if report.filters:
        row_start = len(info_rows) + 5
        ws_info.cell(row=row_start, column=1, value="Filtros aplicados").font = INFO_FONT_TITLE
        for i, (k, v) in enumerate(report.filters.items(), row_start + 1):
            ws_info.cell(row=i, column=1, value=k).font = INFO_LABEL_FONT
            ws_info.cell(row=i, column=2, value=v).font = INFO_FONT

    # Guardar en buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


with st.spinner("Generando Excel..."):
    excel_bytes = _build_excel(report)

filename = f"{report.report_code or uploaded.name.rsplit('.', 1)[0]}.xlsx"

st.download_button(
    label="⬇️ Descargar Excel",
    data=excel_bytes,
    file_name=filename,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary",
    use_container_width=True,
)

st.divider()
st.caption(
    "Los números se convierten automáticamente: separador de miles `,` eliminado, "
    "negativos con guión final `6.000-` → `-6.0`."
)
